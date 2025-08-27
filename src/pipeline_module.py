import io
import pandas as pd
import src.modules as modules
# Ajout des imports nécessaires pour la génération Excel
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime

# === Pipeline principal ===
def pipeline_predict_texts_pil(pil_image):
    # 1. Crop
    cropped_imgs = modules.crop_card(pil_image)
    if not cropped_imgs:
        return None
    cropped_img = cropped_imgs[0]
    # 2. Deskew
    deskewed_img, angle, conf = modules.deskew_image(cropped_img)
    if deskewed_img is None:
        return None
    # 3. Classify
    card_class = modules.classify_card(cropped_img)
    # 4. Extract text boxes
    text_boxes = modules.rec_text(deskewed_img)
    if not text_boxes:
        return None
    # 5. Recognize text
    fields = modules.ocr_text(text_boxes)
    # Ajoute la classe prédite au dictionnaire des champs
    fields["card_class"] = card_class
    return fields

# === Traitement batch pour Streamlit ===
def process_images_to_excel_pil(pil_images):
    results = []
    for pil_image in pil_images:
        infos = pipeline_predict_texts_pil(pil_image)
        if infos is not None and all(x is not None and x != '' for k, x in infos.items() if k != "card_class"):
            results.append(infos)
    if results:
        df = pd.DataFrame(results)
        return df, len(pil_images), len(results), len(pil_images) - len(results)
    else:
        return None, len(pil_images), 0, len(pil_images)

# === Génération du fichier de demande de paiement ===
def gen_demande_paiement(df, nom_activite, output_xlsx="paiement.xlsx"):
    """
    Génère un fichier Excel de demande de paiement à partir d'un DataFrame et d'un titre d'activité.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Font definitions
    header_font = Font(name="aparijata", size=26, bold=True)
    label_font = Font(name="aparijata", size=12, bold=True)
    value_font = Font(name="aparijata", size=12)
    table_header_font = Font(name="aparijata", size=12, bold=True)
    table_font = Font(name="aparijata", size=12)
    signature_font = Font(name="aparijata", size=16, bold=True)

    # Set column widths
    ws.column_dimensions["A"].width = 9
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    # 1. Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = "DEMANDE DE PAIEMENT ELECTRONIQUE"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # 3. Date row
    ws["A3"] = "DATE :"
    ws["A3"].font = label_font
    ws["B3"] = datetime.today().strftime("%d/%m/%Y")
    ws["B3"].font = value_font
    ws.row_dimensions[3].height = 15

    # 4. Operateur row
    ws["A4"] = "OPERATEUR CHOISI :"
    ws["A4"].font = label_font
    ws["B4"] = "MTN"
    ws["B4"].font = value_font
    ws.row_dimensions[4].height = 15

    # 5. Motif row
    ws["A5"] = "MOTIF :"
    ws["A5"].font = label_font
    ws["B5"] = nom_activite
    ws["B5"].font = value_font
    ws.row_dimensions[5].height = 15

    # 7. Table headers (row 7)
    table_headers = [
        "N°",
        "NOM ET PRENOMS",
        "N° TELEPHONE",
        "PERDIEM",
        "FRAIS DE RETRAIT",
        "TOTAL A PAYER",
    ]
    for col, header in enumerate(table_headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = table_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # == Write the data into the Excel file ==
    # 1: Add numbered column to the DataFrame
    df = df.copy()
    df.insert(0, "N°", range(1, len(df) + 1))

    # 2 : merge columns "nom" et "prenom" en "NOM ET PRENOMS" (ou utiliser NOM_PRENOMS si déjà présent)
    if "NOM_PRENOMS" in df.columns:
        df["NOM ET PRENOMS"] = df["NOM_PRENOMS"]
    else:
        df["NOM ET PRENOMS"] = df["nom"] + " " + df["prenom"]

    # 3. Data rows (remplir uniquement les deux premières colonnes)
    for idx, row in enumerate(df[["N°", "NOM ET PRENOMS"]].itertuples(index=False), start=8):
        ws.cell(row=idx, column=1, value=row[0]).font = table_font
        ws.cell(row=idx, column=2, value=row[1]).font = table_font
        ws.row_dimensions[idx].height = 15

    # 4. Total row
    total_row = 8 + len(df)
    ws.cell(row=total_row, column=3, value="TOTAL").font = table_header_font
    ws.row_dimensions[total_row].height = 15

    # 5. Signature row
    signature_row = total_row + 2
    ws.cell(row=signature_row, column=2, value="Le Demandeur").font = signature_font
    ws.cell(row=signature_row, column=5, value="Le Superviseur").font = signature_font

    # 6. Add grid to the table area
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(
        min_row=7,
        max_row=total_row,
        min_col=1,
        max_col=6,
    ):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    wb.save(output_xlsx)
    print(f"File saved as {output_xlsx}")