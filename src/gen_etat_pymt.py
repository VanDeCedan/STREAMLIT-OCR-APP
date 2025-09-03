from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from num2words import num2words
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image as XLImage
import pandas as pd
import os

# === Génération du dataframe état de paiement ===
def create_etat_paiement_xlsx(df, output_file, titre_activite=None, date_activite=None, lieu_deroulement=None, date_paiement=None):
    # Get final data frame
    df['LIEU PROVENANCE'] = ""
    df["TITRE"] = ""
    df["N° TELEPHONE"]= ""
    # rename columns
    df = df.rename(columns={
        'card_class': 'TYPE DE PIECE',
        'numero_carte': "N° PIECE D'IDENTITE",
        "date_expiration": "DATE D'EXPIRATION"
    })
    
    # 2 : merge columns "NOM" and "PRENOMS" into a new column "NOM_PRENOMS"
    df["NOM ET PRENOMS"] = df["nom"] + " " + df["prenom"]
    df = df.drop(columns=["nom", "prenom"])

    # payement colu
    df['NB FORFAIT ALLER'] = ""
    df['TAUX FORFAIT ALLER'] = ""
    df['NB PERDIEM ALLER RETOUR'] = ""
    df['TAUX FORFAIT ALLER RETOUR'] = ""
    df['NB PERDIEM JOUR COMPLET'] = ""
    df['TAUX PERDIEM JOUR COMPLET'] = ""
    df['MONTANT'] = ""
    df['TRANSPORT'] = ""
    df['FRAIS DE RETRAIT'] = ""
    df['STATUT'] = "PAYE"
    df['MONTANT TOTAL'] ="" 

    # # si pièce d'ideentité = CARTE BIOMETRIQUE renvoyé CB, si CARTE CIP = CIP, si PASSEPORT = PP
    df['TYPE DE PIECE'] = df['TYPE DE PIECE'].replace({
        'CARTE BIOMETRIQUE': 'CB',
        'CARTE CIP': 'CIP',
        'PASSEPORT': 'PP'
    })

    # Change columns formats
    df["No PIECE D'IDENTITE"] = (df["TYPE DE PIECE"] + " " + df["N° PIECE D'IDENTITE"].astype("str") +
                                  " " + "\nEXP :" + " " + 
                                  df["DATE D'EXPIRATION"]
    )   

    

    df = df[["NOM ET PRENOMS", "TITRE", "LIEU PROVENANCE", 
            "NB FORFAIT ALLER", "TAUX FORFAIT ALLER",
            "NB PERDIEM ALLER RETOUR", "TAUX FORFAIT ALLER RETOUR",
            "NB PERDIEM JOUR COMPLET", "TAUX PERDIEM JOUR COMPLET",
            "MONTANT", "TRANSPORT", "FRAIS DE RETRAIT",
            "MONTANT TOTAL", "No PIECE D'IDENTITE" , "N° TELEPHONE", "STATUT",
            ]]

    df_final = df.copy()
    df_final = df_final.reset_index(drop=True)

    # === Create Excel file ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Etat de Paiement"

    # Load the image
    # Correction du chemin du logo pour garantir qu'il soit trouvé
    image_path = os.path.join(os.path.dirname(__file__), "Logo_ABMS.png")

    if os.path.exists(image_path):
        img = XLImage(image_path)
        # resize image
        img.width = 100  # Set the width of the image
        img.height = 100  # Set the height of the image
        ws.add_image(img, 'B1')  # Place the image at cell B1

    # Set up font styles
    font_8 = Font(name="Aparajita", size=8)
    font_14 = Font(name="Aparajita", size=14)
    font_10_bold = Font(name="Aparajita", size=10, bold=True)
    font_12_bold = Font(name="Aparajita", size=12, bold=True)
    font_24 = Font(name="Aparajita", size=24)

    # Set up alignment
    align_right = Alignment(horizontal='right')
    align_left = Alignment(horizontal='left')
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_accross = Alignment(horizontal="centerContinuous", vertical="center", wrap_text=True)

    # Set up borders
    border_thin_all = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
        )
    
    border_thin_sans_bas = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        )
    
    border_thin_sans_haut = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        bottom=Side(style='thin')
        )

    # Gray fill for certain cells
    gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    # write header

    ## 1. Write initial content with font size 8
    ws['O1'] = "Taux de perdiem payé en FCFA"
    ws['O1'].font = font_8

    ws['N2'] = "FORFAIT PERDIEM"
    ws['N2'].font = font_8

    ws['N3'] = "FORFAIT PERDIEM ALLER RETOUR SIMPLE"
    ws['N3'].font = font_8

    # === Create Excel file with multiple sheets (5 persons per sheet) ===
    wb = Workbook()
    max_per_sheet = 5
    num_sheets = (len(df_final) + max_per_sheet - 1) // max_per_sheet
    for sheet_idx in range(num_sheets):
        if sheet_idx == 0:
            ws = wb.active
            ws.title = f"Etat {sheet_idx+1}"
        else:
            ws = wb.create_sheet(title=f"Etat {sheet_idx+1}")

        # Load the image
        image_path = os.path.join(os.path.dirname(__file__), "Logo_ABMS.png")
        if os.path.exists(image_path):
            img = XLImage(image_path)
            img.width = 100
            img.height = 100
            ws.add_image(img, 'B1')

        # Set up font styles
        font_8 = Font(name="Aparajita", size=8)
        font_14 = Font(name="Aparajita", size=14)
        font_10_bold = Font(name="Aparajita", size=10, bold=True)
        font_12_bold = Font(name="Aparajita", size=12, bold=True)
        font_24 = Font(name="Aparajita", size=24)

        # Set up alignment
        align_right = Alignment(horizontal='right')
        align_left = Alignment(horizontal='left')
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_accross = Alignment(horizontal="centerContinuous", vertical="center", wrap_text=True)

        # Set up borders
        border_thin_all = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        border_thin_sans_bas = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
        )
        border_thin_sans_haut = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            bottom=Side(style='thin')
        )
        gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

        # write header
        ws['O1'] = "Taux de perdiem payé en FCFA"
        ws['O1'].font = font_8
        ws['N2'] = "FORFAIT PERDIEM"
        ws['N2'].font = font_8
        ws['N3'] = "FORFAIT PERDIEM ALLER RETOUR SIMPLE"
        ws['N3'].font = font_8
        ws['N4'] = "FORFAIT PERDIEM PREMIER & DERNIER JOUR"
        ws['N4'].font = font_8
        ws['N5'] = "FORFAIT PERDIEM JOUR COMPLET"
        ws['N5'].font = font_8
        ws['N6'] = "Frais de déplacement payés selon les tarifs UNACOB"
        ws['N6'].font = font_8
        ws['Q2'] = "MONTANT"
        ws['Q2'].font = font_8
        ws['Q3'] = "6750 // 11250"
        ws['Q3'].font = font_8
        ws['Q4'] = "47500"
        ws['Q4'].font = font_8
        ws['Q5'] = "40000"
        ws['Q5'].font = font_8
        ws['P6'] = "Carburant remboursé au prix de la pompe pour 16L au 100 Km"
        ws['P6'].font = font_8

        border_ranges = ['N2:P2', 'N3:P3', 'N4:P4', 'N5:P5']
        for cell_range in border_ranges:
            for row in ws[cell_range]:
                for cell in row:
                    if cell.column == 14:
                        cell.border = Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    elif cell.column == 16:
                        cell.border = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    else:
                        cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws['Q2:Q5']:
            for cell in row:
                cell.border = border_thin_all
        for row_num in range(1, 8):
            ws.row_dimensions[row_num].height = 12
        ws['B8'] = "TITRE DE L'ACTIVITE :"
        ws['B8'].font = font_12_bold
        ws['B8'].alignment = align_right
        ws['C8'] = titre_activite if titre_activite else ""
        ws['C8'].font = font_12_bold
        ws['C8'].alignment = align_left
        ws['B10'] = "DATE DE L'ACTIVITE :"
        ws['B10'].font = font_12_bold
        ws['B10'].alignment = align_right
        ws['C10'] = date_activite if date_activite else ""
        ws['C10'].font = font_12_bold
        ws['C10'].alignment = align_left
        ws['I10'] = "LIEU DE DEROULEMENT:"
        ws['I10'].font = font_12_bold
        ws['I10'].alignment = align_right
        ws['J10'] = lieu_deroulement if lieu_deroulement else ""
        ws['J10'].font = font_12_bold
        ws['J10'].alignment = align_left
        ws['O10'] = "DATE DE PAIEMENT :"
        ws['O10'].font = font_12_bold
        ws['O10'].alignment = align_right
        ws['P10'] = date_paiement if date_paiement else ""
        ws['P10'].font = font_12_bold
        ws['P10'].alignment = align_left
        ws['I12'] = "ETAT DE PAIEMENT"
        ws['I12'].font = font_12_bold
        ws['I12'].alignment = align_left
        ws.row_dimensions[8].height = 32
        ws['B8'].alignment = Alignment(horizontal='right', wrap_text=True)
        ws.row_dimensions[9].height = 10
        ws.row_dimensions[10].height = 20
        ws.row_dimensions[11].height = 15
        ws.row_dimensions[12].height = 20

        headers_merged_in_rows = [
            ('A', 'N°'),
            ('B', 'Nom et Prénoms'),
            ('C', 'Titre'),
            ('D', 'Lieu de \nprovenance'),
            ('K', 'Montant \n(A)'),
            ('L', 'Transport \naller-retour \n(B)'),
            ('M', 'Frais de retrait \n(si paiement mobile money) \n(C)'),
            ('N', 'Montant total \n(A+B+C)'),
            ('O', "No pièce d'identité"),
            ('P', 'N° téléphone'),
            ('Q', 'Signature')
        ]
        headers_merged_in_cols = [
            ('E', 'FORFAIT PERDIEM \nALLER RETOUR SIMPLE'),
            ('G', 'FORFAIT PERDIEM 1er \nET DERNIER JOUR DE \nMISSION (11250+25000+11250)'),
            ('I', 'FORFAIT PERDIEM \nJOUR COMPLET'),
        ]
        for col, header in headers_merged_in_rows:
            cell = ws[f"{col}14"]
            cell.value = header
            cell.font = font_10_bold
            cell.alignment = align_center
            cell.border = border_thin_sans_bas
        for col, header in headers_merged_in_cols:
            cell = ws[f"{col}14"]
            cell.value = header
            cell.font = font_10_bold
            cell.alignment = align_accross
            cell.border = Border(
                left=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            next_col = chr(ord(col) + 1)
            ws[f"{next_col}14"].border = Border(
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        ws["E15"] = "Nbre"
        ws["E15"].border = border_thin_all
        ws["E15"].alignment = align_center
        ws["E15"].font = font_10_bold
        ws["F15"] = "Taux Forfait"
        ws["F15"].border = border_thin_all
        ws["F15"].alignment = align_center
        ws["F15"].font = font_10_bold
        ws["G15"] = "Nbre"
        ws["G15"].border = border_thin_all
        ws["G15"].alignment = align_center
        ws["G15"].font = font_10_bold
        ws["H15"] = "Taux Forfait"
        ws["H15"].border = border_thin_all
        ws["H15"].alignment = align_center
        ws["H15"].font = font_10_bold
        ws["I15"] = "Nbre"
        ws["I15"].border = border_thin_all
        ws["I15"].alignment = align_center
        ws["I15"].font = font_10_bold
        ws["J15"] = "Taux Forfait"
        ws["J15"].border = border_thin_all
        ws["J15"].alignment = align_center
        ws["J15"].font = font_10_bold
        empty_rows_cells = ['A', 'B', 'C', 'D', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
        for col in empty_rows_cells:
            ws[f'{col}15'].border = border_thin_sans_haut
        empty_cols_cells = ['F','H','J']
        for col in empty_cols_cells:
            ws[f'{col}14'].border = border_thin_all
            ws[f'{col}15'].border = border_thin_all
        ws.row_dimensions[14].height = 90
        ws.row_dimensions[15].height = 30
        ws.print_title_rows = '1:15'

        # Write data for this sheet (max 5 rows)
        start_idx = sheet_idx * max_per_sheet
        end_idx = min((sheet_idx + 1) * max_per_sheet, len(df_final))
        df_sheet = df_final.iloc[start_idx:end_idx]
        rows_needed = max_per_sheet - len(df_sheet)
        if rows_needed > 0:
            empty_rows = pd.DataFrame([{}] * rows_needed, columns=df_final.columns)
            df_sheet = pd.concat([df_sheet, empty_rows], ignore_index=True)
        start_block_row = 16
        for row_num_in_batch, (_, row) in enumerate(df_sheet.iterrows()):
            current_row = start_block_row + row_num_in_batch
            ws[f'A{current_row}'] = row_num_in_batch + 1
            ws[f'B{current_row}'] = row.get('NOM ET PRENOMS', '') or ""
            ws[f'C{current_row}'] = row.get('TITRE', '') or ""
            ws[f'D{current_row}'] = row.get('LIEU PROVENANCE', '') or ""
            ws[f'E{current_row}'] = row.get('NB FORFAIT ALLER', '') or ""
            ws[f'F{current_row}'] = row.get('TAUX FORFAIT ALLER', '') or ""
            ws[f'G{current_row}'] = row.get('NB PERDIEM ALLER RETOUR', '') or ""
            ws[f'H{current_row}'] = row.get('TAUX FORFAIT ALLER RETOUR', '') or ""
            ws[f'I{current_row}'] = row.get('NB PERDIEM JOUR COMPLET', '') or ""
            ws[f'J{current_row}'] = row.get('TAUX PERDIEM JOUR COMPLET', '') or ""
            ws[f'K{current_row}'] = row.get('MONTANT', '') or ""
            ws[f'L{current_row}'] = row.get('TRANSPORT', '') or ""
            ws[f'M{current_row}'] = row.get('FRAIS DE RETRAIT', '') or ""
            ws[f'N{current_row}'] = row.get('MONTANT TOTAL', '') or ""
            ws[f'O{current_row}'] = row.get("No PIECE D'IDENTITE", '') or ""
            ws[f'P{current_row}'] = row.get('N° TELEPHONE', '') or ""
            ws[f'Q{current_row}'] = row.get("STATUT") or ""
            for col in 'ABCDEFGHIJKLMNOPQ':
                ws[f'{col}{current_row}'].alignment = align_center
                ws[f'{col}{current_row}'].font = font_14
            ws.row_dimensions[current_row].height = 45
            for col in 'ABCDEFGHIJKLMNOPQ':
                ws[f'{col}{current_row}'].border = border_thin_all
        total_row = start_block_row + max_per_sheet
        ws[f'B{total_row}'] = "Total"
        Montant_total = 0
        ws[f'N{total_row}'] = Montant_total
        ws[f'N{total_row}'].alignment = align_center
        ws[f'B{total_row}'].alignment = align_center
        ws[f'B{total_row}'].border = border_thin_all
        ws[f'N{total_row}'].border = border_thin_all
        ws.row_dimensions[total_row].height = 28
        gray_cells = ['A', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'O', 'P','Q']
        for col in gray_cells:
            ws[f'{col}{total_row}'].fill = gray_fill
            ws[f'{col}{total_row}'].border = border_thin_all
        summary_row = total_row + 2
        ws[f'B{summary_row}'] = "Arrêté le présent état à la somme de :"
        ws[f'B{summary_row}'].font = font_24
        montant_total_words = num2words(Montant_total, lang='fr')
        ws[f'G{summary_row}'] = montant_total_words.upper() + " FRANCS CFA"
        ws[f'G{summary_row}'].font = font_24
        column_widths = {
            'A': 2,
            'B': 30,
            'C': 10,
            'D': 10,
            'E': 8,
            'F': 10,
            'G': 8,
            'H': 10,
            'I': 8,
            'J': 10,
            'K': 10,
            'L': 10,
            'M': 10,
            'N': 14,
            'O': 20,
            'P': 15,
            'Q': 10
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.05, header=0.3, footer=0.01)
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
    wb.save(output_file)
    print(f"Excel file created: {output_file}")